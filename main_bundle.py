import re
import sys
from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = (
    Path(sys.executable).parent
    if getattr(sys, "frozen", False)
    else Path(__file__).parent.parent
)

RPA_PATH = BASE_DIR / "RPA" / "RpaBank_report.txt"
PINDODO_PATH = BASE_DIR / "PINDODO" / "Pindodo_report.txt"
OUTPUT_PATH = BASE_DIR / "Результат" / "reconciliation.xlsx"
LOG_PATH = BASE_DIR / "logs" / "reconciliation.log"

RPA_MERGE_KEYS = ["date", "amount", "currency", "card_number", "terminal_id"]
PINDODO_MERGE_KEYS = [
    "transaction_date",
    "amount",
    "currency",
    "card_number",
    "terminal_id",
]


def setup_logger() -> None:
    logger.remove()

    logger.add(
        sys.stdout,
        level="INFO",
        format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level:<8}</level> | <cyan>{module}</cyan> - <white>{message}</white>",
        colorize=True,
    )

    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    logger.add(
        LOG_PATH,
        level="DEBUG",
        format="{time:YYYY-MM-DD HH:mm:ss} | {level:<8} | {module}:{function}:{line} - {message}",
        rotation="10 MB",
        retention="7 days",
        encoding="utf-8",
    )


_RPA_RE = re.compile(
    r"^\s*(\d+)\s+(\d+)\s+"
    r"(\d{8})id\S+\s+"
    r"([\d.,]+)([A-Z]{3})"
    r"(\d{6})\s+(\S+)"
)

_PINDODO_FIELDS = {
    "Local Transaction Date and Time": "local_datetime",
    "Transaction Date": "transaction_date",
    "Transaction Amount": "amount",
    "Transaction Currency": "currency",
    "Retrieval Reference Number": "card_number",
    "Card Acceptor Terminal ID": "terminal_id",
}

_PINDODO_RE = re.compile(r"^\s{2,}([A-Za-z][\w\s]+?)\s{2,}(\S.*?)\s*$")


def parse_rpabank(path: Path) -> pd.DataFrame:
    logger.info(f"Парсим RpaBank: {path}")
    rows = []

    with open(path, encoding="utf-8") as f:
        for line in f:
            m = _RPA_RE.match(line)
            if m:
                rows.append(m.groups())
            else:
                logger.debug(f"Строка не распознана: {line.rstrip()}")

    if not rows:
        logger.error("RpaBank: не найдено ни одной записи")
        raise ValueError(f"Файл пустой или формат не распознан: {path}")

    df = pd.DataFrame(
        rows,
        columns=[
            "number",
            "index",
            "date",
            "amount",
            "currency",
            "card_number",
            "terminal_id",
        ],
    )
    df["date"] = pd.to_datetime(df["date"], format="%Y%m%d").dt.date
    df["amount"] = df["amount"].str.replace(",", ".").astype(float)
    df["card_number"] = df["card_number"].str.zfill(6)

    logger.info(f"RpaBank: загружено {len(df)} записей")
    logger.debug(f"RpaBank: валюты — {df['currency'].value_counts().to_dict()}")
    return df


def parse_pindodo(path: Path) -> pd.DataFrame:
    logger.info(f"Парсим Pindodo: {path}")
    rows, current = [], {}
    columns = list(_PINDODO_FIELDS.values())
    skipped = 0

    def flush(d: dict):
        nonlocal skipped
        if set(columns).issubset(d):
            rows.append([d[c] for c in columns])
        elif d:
            skipped += 1
            logger.debug(
                f"Pindodo: блок пропущен, не хватает полей: {set(columns) - set(d)}"
            )

    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\r\n")
            if re.match(r"^\s*-{10,}\s*$", line):
                flush(current)
                current = {}
                continue
            m = _PINDODO_RE.match(line)
            if m:
                key = m.group(1).strip()
                if key in _PINDODO_FIELDS:
                    current[_PINDODO_FIELDS[key]] = m.group(2).strip()

    flush(current)

    if not rows:
        logger.error("Pindodo: не найдено ни одной записи")
        raise ValueError(f"Файл пустой или формат не распознан: {path}")

    if skipped:
        logger.warning(f"Pindodo: пропущено {skipped} неполных блоков")

    df = pd.DataFrame(rows, columns=columns)
    df["local_datetime"] = pd.to_datetime(df["local_datetime"], format="%Y%m%d%H%M%S")
    df["transaction_date"] = pd.to_datetime(
        df["transaction_date"], format="%Y-%m-%d"
    ).dt.date
    df["amount"] = df["amount"].str.replace(",", ".").astype(float)
    df["card_number"] = df["card_number"].str.zfill(6)

    logger.info(f"Pindodo: загружено {len(df)} записей")
    logger.debug(f"Pindodo: валюты — {df['currency'].value_counts().to_dict()}")
    return df


def reconcile(
    df_rpa: pd.DataFrame,
    df_pin: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    logger.info(f"Начинаем сверку: RpaBank={len(df_rpa)}, Pindodo={len(df_pin)}")

    merged = df_rpa.merge(
        df_pin,
        left_on=RPA_MERGE_KEYS,
        right_on=PINDODO_MERGE_KEYS,
        how="outer",
        suffixes=("_rpa", "_pin"),
        indicator=True,
    )

    def _filter(flag: str) -> pd.DataFrame:
        result = (
            merged[merged["_merge"] == flag]
            .drop(columns="_merge")
            .reset_index(drop=True)
        )
        return pd.DataFrame(result)

    success = _filter("both")
    rpa_only = _filter("left_only")
    pind_only = _filter("right_only")

    total = len(success) + len(rpa_only) + len(pind_only)
    match_pct = len(success) / total * 100 if total else 0

    logger.info(f"✓ Успешные:         {len(success)} ({match_pct:.1f}%)")
    logger.info(f"✗ Только RpaBank:   {len(rpa_only)}")
    logger.info(f"✗ Только Pindodo:   {len(pind_only)}")

    if len(rpa_only):
        logger.debug(f"RpaBank-неуспешные RRN: {rpa_only['card_number'].tolist()}")
    if len(pind_only):
        logger.debug(f"Pindodo-неуспешные RRN: {pind_only['card_number'].tolist()}")

    return success, rpa_only, pind_only


FILLS = {
    "Успешные": PatternFill("solid", start_color="1F7A4D"),
    "RpaBank_неуспешные": PatternFill("solid", start_color="C0392B"),
    "Pindodo_неуспешные": PatternFill("solid", start_color="E67E22"),
}


def _style_sheet(ws, fill: PatternFill, row_count: int) -> None:
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDER

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            width + 4, 40
        )

    last = ws.max_row + 2
    ws.cell(row=last, column=1, value="Итого записей:").font = Font(
        name="Arial", bold=True, size=10
    )
    ws.cell(row=last, column=2, value=row_count).font = Font(
        name="Arial", bold=True, size=10
    )

    ws.freeze_panes = "A2"


def export_to_excel(
    df_success: pd.DataFrame,
    df_rpa_only: pd.DataFrame,
    df_pind_only: pd.DataFrame,
    output_path: Path,
) -> None:
    logger.info(f"Записываем Excel: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sheets = {
        "Успешные": df_success,
        "RpaBank_неуспешные": df_rpa_only,
        "Pindodo_неуспешные": df_pind_only,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.debug(f"Лист '{sheet_name}': {len(df)} строк записано")

    wb = load_workbook(output_path)
    for (sheet_name, df), fill in zip(sheets.items(), FILLS.values()):
        _style_sheet(wb[sheet_name], fill, len(df))

    wb.save(output_path)
    logger.info(f"Excel сохранён: {output_path}")


if __name__ == "__main__":
    setup_logger()
    logger.info("=== Запуск сверки ===")

    try:
        df_rpa = parse_rpabank(RPA_PATH)
        df_pin = parse_pindodo(PINDODO_PATH)

        success, rpa_only, pind_only = reconcile(df_rpa, df_pin)

        export_to_excel(success, rpa_only, pind_only, OUTPUT_PATH)

        logger.info("=== Сверка завершена успешно ===")

    except FileNotFoundError as e:
        logger.error(f"Файл не найден: {e}")
        input("Нажми Enter для выхода...")
        sys.exit(1)
    except Exception as e:
        logger.exception(f"Неожиданная ошибка: {e}")
        input("Нажми Enter для выхода...")
        sys.exit(1)
