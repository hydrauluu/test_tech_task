from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FILLS = {
    "Успешные": PatternFill("solid", start_color="1F7A4D"),
    "RpaBank_неуспешные": PatternFill("solid", start_color="C0392B"),
    "Pindodo_неуспешные": PatternFill("solid", start_color="E67E22"),
}


def _style_sheet(ws, fill: PatternFill) -> None:
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDER

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            width + 4, 40
        )

    last = ws.max_row + 2
    ws.cell(row=last, column=1, value="Итого записей:").font = Font(
        name="Arial", bold=True, size=10
    )
    ws.cell(row=last, column=2, value=f"=COUNTA(A2:A{last - 2})").font = Font(
        name="Arial", bold=True, size=10
    )

    ws.freeze_panes = "A2"


def export_to_excel(
    df_success: pd.DataFrame,
    df_rpa_only: pd.DataFrame,
    df_pind_only: pd.DataFrame,
    output_path: str | Path,
) -> None:
    """
    Записывает три DataFrame на отдельные листы Excel:
        Успешные            — зелёный заголовок
        RpaBank_неуспешные  — красный заголовок
        Pindodo_неуспешные  — оранжевый заголовок
    """
    sheets = {
        "Успешные": df_success,
        "RpaBank_неуспешные": df_rpa_only,
        "Pindodo_неуспешные": df_pind_only,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    for sheet_name, fill in FILLS.items():
        _style_sheet(wb[sheet_name], fill)

    wb.save(output_path)
    print(f"Сохранено: {output_path}")
